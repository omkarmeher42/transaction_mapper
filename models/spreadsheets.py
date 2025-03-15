import openpyxl
from openpyxl.styles import Alignment
from datetime import datetime

class SpreadSheet:
    def __init__(self, sheet_name, user):
        self.sheet_name = sheet_name
        self.user = user

    def create_sheet(self):
        workbook = openpyxl.Workbook()
        workbook.save(f'{self.sheet_name}.xlsx')
        return workbook
    
    def create_sheet_path(self):
        self.sheet_path = self.user.current_sheet_path

    def is_blank(self):
        self.create_sheet_path()
        workbook = openpyxl.load_workbook(self.sheet_path)
        sheet = workbook.active
        # Check if all cells in the sheet are empty
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    return False
        return True

    def apply_template(self):
        self.create_sheet_path()
        workbook = openpyxl.load_workbook(self.sheet_path)
        sheet = workbook.active

        # Define the headings that match the input form fields in transactions.html
        headings = ["Sr No", "Date", "Transaction Title", "Amount", "Category", "Sub Category", "Payment Method"]

        # Insert 2 rows at the top
        sheet.insert_rows(1, 2)

        # Merge the columns for the file name
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headings))

        # Add the file name to the merged cells
        sheet.cell(row=1, column=1, value=self.sheet_name)

        # Add the headings to the third row
        for col_num, heading in enumerate(headings, 1):
            cell = sheet.cell(row=3, column=col_num, value=heading)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Center-align all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Resize columns to fit the length of their column heading
        for col in sheet.iter_cols(min_row=3, max_row=3):
            max_length = len(col[0].value)
            column = col[0].column_letter  # Get the column name
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Auto-increment "Sr No" column
        for row_num in range(4, sheet.max_row + 1):
            sheet.cell(row=row_num, column=1, value=row_num - 3)

        # Merge cells for "Total Amount" and add the value
        sheet.merge_cells(start_row=1, start_column=len(headings) + 1, end_row=2, end_column=len(headings) + 2)
        sheet.cell(row=1, column=len(headings) + 1, value="Total Amount").alignment = Alignment(horizontal='center', vertical='center')

        # Merge cells for the sum of the Amount column
        max_row = sheet.max_row
        sheet.merge_cells(start_row=3, start_column=len(headings) + 1, end_row=4, end_column=len(headings) + 2)
        amount_column = 'D'  # Assuming "Amount" is in column D
        sheet.cell(row=3, column=len(headings) + 1, value=f"=SUM({amount_column}4:{amount_column}{max_row})").alignment = Alignment(horizontal='center', vertical='center')

        workbook.save(self.sheet_path)

    def get_current_date(self):
        return datetime.now().strftime("%Y-%m-%d")
