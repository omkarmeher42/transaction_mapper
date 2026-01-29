from models.users import User, db
from services.user_services import UserService
import os
from flask import send_file, abort
import logging
import openpyxl
from openpyxl.styles import Alignment
from services.email_service import send_email  # Import the send_email function
import io
from datetime import datetime, timedelta, date
from sqlalchemy import func, extract
import statistics

class TransactionServices:

    @staticmethod
    def generate_file_content(month, year, user_id):
        try:
            if not user_id:
                logging.error("User ID is missing")
                return None, None, None

            logging.debug(f"Starting download for month: {month}, year: {year}, user_id: {user_id}")

            user = UserService.get_user_by_id(user_id)
            if not user:
                logging.error(f"No user found with id: {user_id}")
                return None, None, None

            file_name = f'{month}_{year}'
            file_path = f'{os.getcwd()}/Sheets/{user.user_name}/{file_name}.xlsx'

            if os.path.exists(file_path):
                logging.debug(f"XLSX file found: {file_path}")
                response = send_file(file_path, as_attachment=True)
                logging.debug(f"XLSX file sent: {file_path}")
                return file_path, response, file_name
            else:
                logging.error(f"XLSX file not found: {file_path}")
                abort(404, description="File not found")
        except Exception as e:
            logging.error(f"Error in generate_file_content: {e}")
            return None, None, None

    @staticmethod
    def send_file_via_email(file_path, user_id):
        try:
            user = UserService.get_user_by_id(user_id)
            if not user:
                logging.error(f"No user found with id: {user_id}")
                return

            email_address = user.email_id
            subject = f"{os.path.basename(file_path)} - Monthly Expense Report"
            body = "File attached, Monthly expense report."

            print('\n\nSending email to:', email_address)

            # Create BytesIO object from file
            with open(file_path, 'rb') as f:
                file_data = io.BytesIO(f.read())
                file_data.filename = os.path.basename(file_path)

            send_email(email_address, subject, body, file_data)
            logging.info(f"Email sent successfully to {email_address}")
        except Exception as e:
            logging.error(f"Failed to send email to {email_address}: {e}")

    @staticmethod
    def process_recurring_transactions(user_id):
        from models.budget_recurring import RecurringTransaction
        from models.transactions import Transaction
        from models import db
        from datetime import datetime, date
        
        today = date.today()
        # Find active recurring transactions for this user that haven't been logged this month
        # and whose day has passed or is today
        recurring_txs = RecurringTransaction.query.filter_by(user_id=user_id, is_active=True).all()
        
        for rtx in recurring_txs:
            # Check if already logged this month
            if rtx.last_logged and rtx.last_logged.month == today.month and rtx.last_logged.year == today.year:
                continue
                
            # Check if today is matching or past the scheduled day
            if today.day >= rtx.day_of_month:
                try:
                    # Create the transaction
                    new_tx = Transaction(
                        user_id=user_id,
                        date=today,
                        title=f"[Recurring] {rtx.title}",
                        amount=rtx.amount,
                        category=rtx.category,
                        sub_category=rtx.sub_category,
                        payment_method=rtx.payment_method
                    )
                    db.session.add(new_tx)
                    
                    # Update the recurring transaction's last logged date
                    rtx.last_logged = today
                    db.session.commit()
                    logging.info(f"Automatically logged recurring transaction: {rtx.title} for user {user_id}")
                except Exception as e:
                    logging.error(f"Error processing recurring transaction {rtx.id}: {e}")
                    db.session.rollback()

    @staticmethod
    def get_analytics_data(user_id):
        """Generate comprehensive analytics data for dashboard"""
        from models.transactions import Transaction
        
        try:
            now = datetime.now()
            current_month = now.month
            current_year = now.year
            
            # Get current month transactions
            current_month_txs = Transaction.query.filter(
                Transaction.user_id == user_id,
                extract('month', Transaction.date) == current_month,
                extract('year', Transaction.date) == current_year
            ).all()
            
            analytics = {
                'monthly_trend': TransactionServices._calculate_monthly_trend(user_id, current_year),
                'category_growth': TransactionServices._calculate_category_growth(user_id),
                'average_transaction': TransactionServices._calculate_average_transaction(current_month_txs),
                'daily_average': TransactionServices._calculate_daily_average(user_id, current_month, current_year),
                'weekly_pattern': TransactionServices._calculate_weekly_pattern(current_month_txs),
                'highest_spending_day': TransactionServices._get_highest_spending_day(current_month_txs),
                'anomalies': TransactionServices._detect_anomalies(user_id, current_month, current_year),
                'savings_rate': TransactionServices._calculate_savings_rate(user_id, current_month, current_year),
            }
            
            return analytics
        except Exception as e:
            logging.error(f"Error generating analytics data: {e}")
            return {}
    
    @staticmethod
    def _calculate_monthly_trend(user_id, current_year):
        """Calculate month-over-month spending for last 12 months"""
        from models.transactions import Transaction
        
        monthly_totals = {}
        for month in range(1, 13):
            total = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.user_id == user_id,
                extract('month', Transaction.date) == month,
                extract('year', Transaction.date) == current_year
            ).scalar() or 0
            month_name = datetime(current_year, month, 1).strftime('%b')
            monthly_totals[month_name] = float(total)
        
        return monthly_totals
    
    @staticmethod
    def _calculate_category_growth(user_id):
        """Compare this month vs last month spending by category"""
        from models.transactions import Transaction
        
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        
        # Last month
        if current_month == 1:
            last_month, last_year = 12, current_year - 1
        else:
            last_month, last_year = current_month - 1, current_year
        
        # Current month by category
        current_data = db.session.query(
            Transaction.category, func.sum(Transaction.amount)
        ).filter(
            Transaction.user_id == user_id,
            extract('month', Transaction.date) == current_month,
            extract('year', Transaction.date) == current_year
        ).group_by(Transaction.category).all()
        
        # Last month by category
        last_data = db.session.query(
            Transaction.category, func.sum(Transaction.amount)
        ).filter(
            Transaction.user_id == user_id,
            extract('month', Transaction.date) == last_month,
            extract('year', Transaction.date) == last_year
        ).group_by(Transaction.category).all()
        
        current_dict = {cat: float(amt) for cat, amt in current_data}
        last_dict = {cat: float(amt) for cat, amt in last_data}
        
        # Calculate growth
        growth = {}
        all_categories = set(current_dict.keys()) | set(last_dict.keys())
        
        for category in all_categories:
            current_amt = current_dict.get(category, 0)
            last_amt = last_dict.get(category, 0)
            
            if last_amt == 0:
                growth[category] = 100 if current_amt > 0 else 0
            else:
                growth[category] = round(((current_amt - last_amt) / last_amt) * 100, 1)
        
        return growth
    
    @staticmethod
    def _calculate_average_transaction(transactions):
        """Calculate average transaction amount"""
        if not transactions:
            return 0
        return round(sum(tx.amount for tx in transactions) / len(transactions), 2)
    
    @staticmethod
    def _calculate_daily_average(user_id, month, year):
        """Calculate average spending per day"""
        from models.transactions import Transaction
        
        total = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            extract('month', Transaction.date) == month,
            extract('year', Transaction.date) == year
        ).scalar() or 0
        
        # Approximate days in month
        if month == 12:
            next_month_start = date(year + 1, 1, 1)
        else:
            next_month_start = date(year, month + 1, 1)
        
        month_start = date(year, month, 1)
        days_in_month = (next_month_start - month_start).days
        
        return round(float(total) / days_in_month, 2)
    
    @staticmethod
    def _calculate_weekly_pattern(transactions):
        """Analyze spending by day of week"""
        if not transactions:
            return {}
        
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        daily_totals = {day: 0 for day in days}
        daily_counts = {day: 0 for day in days}
        
        for tx in transactions:
            day_name = days[tx.date.weekday()]
            daily_totals[day_name] += tx.amount
            daily_counts[day_name] += 1
        
        # Calculate averages
        pattern = {}
        for day in days:
            if daily_counts[day] > 0:
                pattern[day] = round(daily_totals[day] / daily_counts[day], 2)
            else:
                pattern[day] = 0
        
        return pattern
    
    @staticmethod
    def _get_highest_spending_day(transactions):
        """Find the day with highest total spending"""
        if not transactions:
            return None
        
        daily_totals = {}
        for tx in transactions:
            day = tx.date.strftime('%Y-%m-%d')
            daily_totals[day] = daily_totals.get(day, 0) + tx.amount
        
        if daily_totals:
            highest_day = max(daily_totals, key=daily_totals.get)
            return {
                'date': highest_day,
                'amount': round(daily_totals[highest_day], 2)
            }
        return None
    
    @staticmethod
    def _detect_anomalies(user_id, month, year):
        """Detect unusual spending patterns"""
        from models.transactions import Transaction
        
        try:
            current_txs = Transaction.query.filter(
                Transaction.user_id == user_id,
                extract('month', Transaction.date) == month,
                extract('year', Transaction.date) == year
            ).all()
            
            if not current_txs or len(current_txs) < 5:
                return []
            
            # Get last 3 months average for each category
            anomalies = []
            categories = set(tx.category for tx in current_txs)
            
            for category in categories:
                current_total = sum(tx.amount for tx in current_txs if tx.category == category)
                
                # Get last 3 months data
                three_months_ago = date.today() - timedelta(days=90)
                historical = Transaction.query.filter(
                    Transaction.user_id == user_id,
                    Transaction.category == category,
                    Transaction.date < date(year, month, 1),
                    Transaction.date >= three_months_ago
                ).all()
                
                if historical:
                    avg_historical = sum(tx.amount for tx in historical) / 3  # 3 months
                    
                    # Flag if current is 50% higher than average
                    if current_total > avg_historical * 1.5:
                        anomalies.append({
                            'category': category,
                            'current': round(current_total, 2),
                            'historical_avg': round(avg_historical, 2),
                            'increase_pct': round(((current_total - avg_historical) / avg_historical) * 100, 1)
                        })
            
            return anomalies
        except Exception as e:
            logging.error(f"Error detecting anomalies: {e}")
            return []
    
    @staticmethod
    def _calculate_savings_rate(user_id, month, year):
        """Estimate savings rate (requires income data if available)"""
        from models.transactions import Transaction
        
        # This is a simple calculation based on spending
        # Can be enhanced if income data is available
        current_total = db.session.query(func.sum(Transaction.amount)).filter(
            Transaction.user_id == user_id,
            extract('month', Transaction.date) == month,
            extract('year', Transaction.date) == year
        ).scalar() or 0
        
        # Calculate average spending last 3 months
        three_months_avg = 0
        for i in range(1, 4):
            if month - i <= 0:
                check_month = 12 + month - i
                check_year = year - 1
            else:
                check_month = month - i
                check_year = year
            
            monthly_total = db.session.query(func.sum(Transaction.amount)).filter(
                Transaction.user_id == user_id,
                extract('month', Transaction.date) == check_month,
                extract('year', Transaction.date) == check_year
            ).scalar() or 0
            three_months_avg += monthly_total
        
        three_months_avg = three_months_avg / 3
        
        if three_months_avg == 0:
            savings_comparison = "N/A"
        else:
            diff = ((current_total - three_months_avg) / three_months_avg) * 100
            if diff < 0:
                savings_comparison = f"↓ {abs(round(diff, 1))}% (Good!)"
            else:
                savings_comparison = f"↑ {round(diff, 1)}%"
        
        return {
            'current_month': round(float(current_total), 2),
            'three_month_avg': round(three_months_avg, 2),
            'trend': savings_comparison
        }


class ExcelService:

    @staticmethod
    def append_transaction_data(file_path, transaction_data):
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Find the last row with data (excluding the sum row)
        last_row = 3  # Start after headers
        for row in range(4, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value is not None:
                last_row = row

        # Calculate the next Sr No
        next_sr_no = 1 if last_row == 3 else sheet.cell(row=last_row, column=1).value + 1

        # Append the transaction data to the sheet
        new_row = [
            next_sr_no,
            transaction_data['date'],
            transaction_data['title'],
            float(transaction_data['amount']),
            transaction_data['category'],
            transaction_data['sub_category'],
            transaction_data['payment_method'],
        ]
        sheet.append(new_row)

        # Update the sum formula
        max_row = sheet.max_row
        headings = 7
        amount_column = 'D'  # Assuming "Amount" is in column D
        sheet.cell(row=3, column=headings + 1, value=f"=SUM({amount_column}4:{amount_column}{max_row})").alignment = Alignment(horizontal='center', vertical='center')

        # Center-align all cells in the new row
        for cell in sheet[sheet.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        workbook.save(file_path)
        logging.debug(f"Transaction data appended to the Excel sheet with Sr No: {next_sr_no}")

    @staticmethod
    def update_transaction_data(file_path, transaction_data):
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Find row by transaction_id (Sr No)
        row_num = None
        for row in range(4, sheet.max_row + 1):  # Start from row 4 (data starts after header)
            if str(sheet.cell(row=row, column=1).value) == str(transaction_data['transaction_id']):
                row_num = row
                break

        if row_num is None:
            raise ValueError("Transaction not found")

        # Update the row with new data
        sheet.cell(row=row_num, column=2, value=transaction_data['date'])
        sheet.cell(row=row_num, column=3, value=transaction_data['title'])
        sheet.cell(row=row_num, column=4, value=float(transaction_data['amount']))
        sheet.cell(row=row_num, column=5, value=transaction_data['category'])
        sheet.cell(row=row_num, column=6, value=transaction_data['sub_category'])
        sheet.cell(row=row_num, column=7, value=transaction_data['payment_method'])

        # Center-align all cells in the updated row
        for cell in sheet[row_num]:
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        workbook.save(file_path)

    @staticmethod
    def delete_transaction_data(file_path, transaction_id):
        # Load the workbook and select the active sheet
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Find all non-empty rows and their data
        valid_rows = []
        for row in range(4, sheet.max_row + 1):
            # Check if row has actual data (not just empty cells)
            if any(sheet.cell(row=row, column=col).value for col in range(2, 8)):
                sr_no = sheet.cell(row=row, column=1).value
                if str(sr_no) == str(transaction_id):
                    continue  # Skip the row to be deleted
                row_data = [sheet.cell(row=row, column=col).value for col in range(1, 8)]
                valid_rows.append(row_data)

        # Clear all data rows
        if sheet.max_row > 3:  # Only if there are data rows
            sheet.delete_rows(4, sheet.max_row - 3)

        # Rewrite all valid rows with new Sr No
        for idx, row_data in enumerate(valid_rows, 1):
            new_row = [idx] + row_data[1:]  # Replace old Sr No with new sequential number
            sheet.append(new_row)

            # Center-align all cells in the row
            for cell in sheet[sheet.max_row]:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # Update the sum formula
        max_row = sheet.max_row
        headings = 7
        amount_column = 'D'
        sheet.cell(row=3, column=headings + 1, value=f"=SUM({amount_column}4:{amount_column}{max_row})").alignment = Alignment(horizontal='center', vertical='center')

        # Save the workbook
        workbook.save(file_path)